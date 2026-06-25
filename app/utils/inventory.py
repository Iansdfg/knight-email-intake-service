from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader
from xlrd import open_workbook


@dataclass(frozen=True)
class InventoryMetadata:
    extension: str | None
    page_count: int | None = None
    sheet_count: int | None = None


def inspect_document(filename: str, content: bytes, mime_type: str | None) -> InventoryMetadata:
    extension = Path(filename).suffix.lower().lstrip(".") or None
    page_count = _pdf_page_count(content) if _is_pdf(extension, mime_type) else None
    sheet_count = _excel_sheet_count(content) if _is_excel(extension, mime_type) else None
    return InventoryMetadata(extension=extension, page_count=page_count, sheet_count=sheet_count)


def _is_pdf(extension: str | None, mime_type: str | None) -> bool:
    return extension == "pdf" or mime_type == "application/pdf"


def _is_excel(extension: str | None, mime_type: str | None) -> bool:
    excel_mime_types = {
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return extension in {"xls", "xlsx", "xlsm"} or mime_type in excel_mime_types


def _pdf_page_count(content: bytes) -> int | None:
    try:
        reader = PdfReader(BytesIO(content))
        return len(reader.pages)
    except Exception:
        return None


def _excel_sheet_count(content: bytes) -> int | None:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        try:
            return len(workbook.sheetnames)
        finally:
            workbook.close()
    except Exception:
        try:
            workbook = open_workbook(file_contents=content, on_demand=True)
            return workbook.nsheets
        except Exception:
            return None
