from dataclasses import dataclass
from html import unescape
from html.parser import HTMLParser
from io import BytesIO
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class ParsedTable:
    title: str
    rows: list[list[str]]


class _HTMLTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.tables: list[list[list[str]]] = []
        self._current_table: list[list[str]] | None = None
        self._current_row: list[str] | None = None
        self._current_cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "table":
            self._current_table = []
        elif tag == "tr" and self._current_table is not None:
            self._current_row = []
        elif tag in {"td", "th"} and self._current_row is not None:
            self._current_cell = []

    def handle_data(self, data: str) -> None:
        if self._current_cell is not None:
            self._current_cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag in {"td", "th"} and self._current_cell is not None and self._current_row is not None:
            value = " ".join("".join(self._current_cell).split())
            self._current_row.append(unescape(value))
            self._current_cell = None
        elif tag == "tr" and self._current_row is not None and self._current_table is not None:
            if any(cell for cell in self._current_row):
                self._current_table.append(self._current_row)
            self._current_row = None
        elif tag == "table" and self._current_table is not None:
            if self._current_table:
                self.tables.append(self._current_table)
            self._current_table = None


def email_tables_workbook_content(email_body: str) -> tuple[bytes | None, int]:
    tables = _parse_html_tables(email_body)
    if not tables:
        tables = _parse_text_tables(_html_to_text(email_body))

    if not tables:
        return None, 0

    workbook = Workbook()
    workbook.remove(workbook.active)
    for index, table in enumerate(tables, start=1):
        sheet = workbook.create_sheet(_sheet_title(table.title, index))
        for row in table.rows:
            sheet.append(row)
        _format_sheet(sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), len(tables)


def _parse_html_tables(email_body: str) -> list[ParsedTable]:
    parser = _HTMLTableParser()
    parser.feed(email_body)
    return [
        ParsedTable(title=f"Table {index}", rows=rows)
        for index, rows in enumerate(parser.tables, start=1)
    ]


def _html_to_text(email_body: str) -> str:
    text = re.sub(r"(?i)<br\s*/?>", "\n", email_body)
    text = re.sub(r"(?i)</(p|div|tr|table|pre)>", "\n", text)
    text = re.sub(r"<[^>]+>", " ", text)
    return unescape(text)


def _parse_text_tables(text: str) -> list[ParsedTable]:
    normalized = " ".join(text.split())
    tables = []
    tables.extend(_parse_historical_values(normalized))
    tables.extend(_parse_summary_totals(normalized))
    tables.extend(_parse_auto_liability_losses(normalized))
    return tables


def _parse_historical_values(text: str) -> list[ParsedTable]:
    marker = "Historical Values:"
    next_marker = "SUMMARY TOTALS"
    section = _section_between(text, marker, next_marker)
    if not section:
        return []

    pattern = re.compile(
        r"(?P<year>\d{2}-\d{2})\s+"
        r"(?P<units>[\d,]+)\s+"
        r"(?P<mileage_ppu>[\d,']+)\s+"
        r"(?P<total_mileage>[\d,]+)\s+"
        r"\$?(?P<revenue_ppu>[\d,]+)\s+"
        r"\$?(?P<total_revenue>[\d,]+)"
    )
    rows = [["Year", "Units", "Mileage PPU", "Total Mileage", "Revenue PPU", "Total Revenue"]]
    for match in pattern.finditer(section):
        rows.append([match.group(name) for name in (
            "year",
            "units",
            "mileage_ppu",
            "total_mileage",
            "revenue_ppu",
            "total_revenue",
        )])
    return [ParsedTable("Historical Values", rows)] if len(rows) > 1 else []


def _parse_summary_totals(text: str) -> list[ParsedTable]:
    section = _section_between(text, "SUMMARY TOTALS", "AUTO LIABILITY LOSSES")
    if not section:
        return []

    pattern = re.compile(
        r"(?P<policy_year>\d{4}-\d{4})\s+"
        r"\$?\s*(?P<loss_reserve>[\d,.-]+)\s+"
        r"\$?\s*(?P<total_incurred>[\d,.-]+)\s+"
        r"(?P<total_claims>\d+)\s+"
        r"(?P<units>\d+)\s+"
        r"(?P<loss_frequency>\d+%)\s+"
        r"(?P<mileage_per_unit>[\d,]+)\s+"
        r"(?P<total_mileage>[\d,]+)\s+"
        r"(?P<revenue_per_truck>[\d,]+)"
    )
    rows = [[
        "Policy Year",
        "Loss Reserve",
        "Total Incurred",
        "Total Claims",
        "Units",
        "Loss Frequency",
        "Mileage Per Unit",
        "Total Mileage",
        "Revenue Per Truck",
    ]]
    for match in pattern.finditer(_normalize_currency_dashes(section)):
        rows.append([match.group(name) for name in (
            "policy_year",
            "loss_reserve",
            "total_incurred",
            "total_claims",
            "units",
            "loss_frequency",
            "mileage_per_unit",
            "total_mileage",
            "revenue_per_truck",
        )])
    return [ParsedTable("Summary Totals", rows)] if len(rows) > 1 else []


def _parse_auto_liability_losses(text: str) -> list[ParsedTable]:
    section = _section_after(text, "AUTO LIABILITY LOSSES")
    if not section:
        return []

    pattern = re.compile(
        r"(?P<company>[A-Z][A-Z\s.-]+?)\s+"
        r"(?P<policy_year>\d{4}-\d{4})\s+"
        r"(?P<policy_number>\S+)\s+"
        r"(?P<claims>\d+)\s+"
        r"\$?\s*(?P<loss_reserve>[\d,.-]+)\s+"
        r"\$?\s*(?P<total_incurred>[\d,.-]+)"
    )
    rows = [[
        "Insurance Company",
        "Policy Year",
        "Policy Number",
        "No. of Claims",
        "Loss Reserve",
        "Total Incurred",
    ]]
    for match in pattern.finditer(_normalize_currency_dashes(section)):
        rows.append([match.group(name).strip() for name in (
            "company",
            "policy_year",
            "policy_number",
            "claims",
            "loss_reserve",
            "total_incurred",
        )])
    return [ParsedTable("Auto Liability Losses", rows)] if len(rows) > 1 else []


def _section_between(text: str, start: str, end: str) -> str:
    start_index = text.find(start)
    if start_index == -1:
        return ""
    end_index = text.find(end, start_index + len(start))
    if end_index == -1:
        return text[start_index + len(start):]
    return text[start_index + len(start):end_index]


def _section_after(text: str, start: str) -> str:
    start_index = text.find(start)
    if start_index == -1:
        return ""
    return text[start_index + len(start):]


def _normalize_currency_dashes(value: str) -> str:
    return re.sub(r"\$\s*-", "$0", value)


def _sheet_title(title: str, index: int) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "", title).strip() or f"Table {index}"
    return cleaned[:31]


def _format_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width + 2, 12), 42)
